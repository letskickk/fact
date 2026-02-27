"""Load reference files (PDF/TXT/HWP/HWPX) into text chunks for RAG."""

from __future__ import annotations

import logging
import struct
import zlib
import zipfile
from pathlib import Path
from xml.etree import ElementTree

logger = logging.getLogger(__name__)

FACTS_DIR = Path("data/facts")
CHUNK_SIZE = 800  # characters per chunk
CHUNK_OVERLAP = 200  # overlap between chunks


def _extract_pdf_text(path: Path) -> str:
    """Extract all text from a PDF file using pypdf."""
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    texts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            texts.append(text)
    return "\n".join(texts)


def _extract_txt_text(path: Path) -> str:
    """Read a plain text file."""
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_hwp_text(path: Path) -> str:
    """Extract text from HWP (binary OLE2) file using olefile."""
    import olefile

    ole = olefile.OleFileIO(str(path))
    texts = []

    # BodyText 스트림에서 텍스트 추출
    for stream_path in ole.listdir():
        if stream_path[0] == "BodyText":
            data = ole.openstream(stream_path).read()
            # HWP body는 zlib 압축일 수 있음
            try:
                data = zlib.decompress(data, -15)
            except zlib.error:
                pass

            # 바이너리 레코드에서 텍스트 추출
            i = 0
            while i < len(data) - 4:
                header = struct.unpack_from("<I", data, i)[0]
                rec_type = header & 0x3FF
                rec_len = (header >> 20) & 0xFFF
                i += 4
                if rec_len == 0xFFF:
                    if i + 4 <= len(data):
                        rec_len = struct.unpack_from("<I", data, i)[0]
                        i += 4
                # rec_type 67 = HWPTAG_PARA_TEXT
                if rec_type == 67 and i + rec_len <= len(data):
                    raw = data[i:i + rec_len]
                    # UTF-16LE 텍스트, 제어문자 필터링
                    chars = []
                    for ci in range(0, len(raw) - 1, 2):
                        code = struct.unpack_from("<H", raw, ci)[0]
                        if code >= 32:
                            chars.append(chr(code))
                        elif code in (13, 10):
                            chars.append("\n")
                    line = "".join(chars).strip()
                    if line:
                        texts.append(line)
                i += rec_len

    ole.close()
    return "\n".join(texts)


def _extract_hwpx_text(path: Path) -> str:
    """Extract text from HWPX (ZIP/XML-based) file."""
    texts = []
    with zipfile.ZipFile(str(path)) as z:
        for name in sorted(z.namelist()):
            if "section" in name.lower() and name.endswith(".xml"):
                with z.open(name) as f:
                    tree = ElementTree.parse(f)
                    root = tree.getroot()
                    # 모든 네임스페이스의 텍스트 노드 추출
                    for elem in root.iter():
                        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                        if tag == "t" and elem.text:
                            texts.append(elem.text)
    return "\n".join(texts)


def _extract_xlsx_text(path: Path) -> str:
    """Extract text from Excel (.xlsx) file using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    texts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                texts.append(" | ".join(cells))
    wb.close()
    return "\n".join(texts)


def _chunk_text(text: str, source: str) -> list[dict]:
    """Split text into overlapping chunks with metadata."""
    chunks = []
    # Clean up whitespace
    text = " ".join(text.split())
    if not text:
        return chunks

    start = 0
    idx = 0
    while start < len(text):
        end = start + CHUNK_SIZE
        chunk_text = text[start:end]
        if chunk_text.strip():
            chunks.append({
                "id": f"{source}_{idx}",
                "text": chunk_text.strip(),
                "source": source,
            })
            idx += 1
        start += CHUNK_SIZE - CHUNK_OVERLAP

    return chunks


def load_all_documents() -> list[dict]:
    """Load all documents from data/facts/ and return text chunks.

    Returns list of {"id": str, "text": str, "source": str}.
    """
    if not FACTS_DIR.exists():
        logger.info("No facts directory found at %s", FACTS_DIR)
        return []

    all_chunks = []
    for f in sorted(FACTS_DIR.rglob("*")):
        if f.name.startswith(".") or not f.is_file():
            continue

        try:
            ext = f.suffix.lower()
            if ext == ".pdf":
                text = _extract_pdf_text(f)
            elif ext == ".hwp":
                text = _extract_hwp_text(f)
            elif ext == ".hwpx":
                text = _extract_hwpx_text(f)
            elif ext in (".xlsx", ".xls"):
                text = _extract_xlsx_text(f)
            elif ext in (".txt", ".md", ".csv"):
                text = _extract_txt_text(f)
            else:
                logger.info("Skipping unsupported file: %s", f.name)
                continue

            # 하위폴더 경로 포함한 소스명
            rel_path = f.relative_to(FACTS_DIR)
            source_name = str(rel_path).replace("\\", "/")
            chunks = _chunk_text(text, source_name)
            all_chunks.extend(chunks)
            logger.info("Loaded %s: %d chunks (%d chars)", source_name, len(chunks), len(text))

        except Exception as e:
            logger.warning("Failed to load %s: %s", f.name, e)

    logger.info("Total reference chunks: %d", len(all_chunks))
    return all_chunks
